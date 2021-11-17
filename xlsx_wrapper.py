"""
----------------------
 XLSX handler segment
----------------------
"""

import re
from pathlib import Path
from re import Match
from typing import IO, Iterable, List, Dict, Optional

import numpy as np
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class WrapperXLSX:
    def get_sheet_column_names(self) -> List:
        _col_names = []
        for _col in self._sheet.iter_cols(max_row=1):
            _col_names.append(_col[0].value)
        return _col_names

    def get_rows_data(self, col_amount: int = None, limit: int = None) -> np.array:
        data: List[List] = []

        if col_amount is None:
            col_amount = self.get_sheet_column_names().__len__()

        limit = limit + 1 if limit else None
        for i, row in enumerate(self._sheet.iter_rows(values_only=True,
                                                      max_row=limit,
                                                      max_col=col_amount)):
            if i != 0:
                data.append([col for col in row])

        return np.array(data)

    def __init__(self, sheet: Worksheet):
        if sheet is None:
            raise ValueError('sheet is None')
        self._sheet = sheet


class OpenXLSX:
    _LIMIT_FILE_LEN = 50

    @classmethod
    def _check_url(cls, url: str) -> Match:
        url_match = re.match(r"^(?:(?P<scheme>https?)://)?"
                             r"(?P<domain>[\w.-]+)"
                             r"(?::(?P<port>\d+))?"
                             r"(?P<routing>(?:/(?P<id>[\w.-]+))*)?/?"
                             r"(?P<parameters>\?[^#]+)?"
                             r"(?P<anchor>#.+)?$",
                             url)
        return url_match

    @classmethod
    def create_by_cached_file(cls, url: str, path_to_place: str = 'cache\\') -> Optional["OpenXLSX"]:
        # check url:
        url_match = cls._check_url(url)
        if not url_match:
            raise ValueError(f'url "{url}" has wrong format')

        # check path_to_place
        if not Path(path_to_place).is_dir():
            raise ValueError(f'path_to_place "{path_to_place}" is not a dictionary')

        id_ = url_match.group('id')
        file_name = path_to_place + str(id_)[:cls._LIMIT_FILE_LEN] + '.xlsx'

        return cls(file_name) if Path(file_name).is_file() else None

    @classmethod
    def create_by_download_from_google_sheets(cls, url: str, path_to_place: str = 'cache\\') -> "OpenXLSX":
        """
        :param path_to_place: a path where to a downloaded file will be placed
        :param url: there is a file to download
        :return: file full name to access it
        """
        import requests

        # check url:
        url_match = cls._check_url(url)
        if not url_match:
            raise ValueError(f'url "{url}" has wrong format')

        # check path_to_place
        if not Path(path_to_place).is_dir():
            raise ValueError(f'path_to_place "{path_to_place}" is not a dictionary')

        id_ = url_match.group('id')
        actual_file_url = url if url.endswith('/') else url + '/'

        download_file_request = '{}export?format=xlsx&id={}' \
            .format(actual_file_url, id_)

        with requests.get(download_file_request) as response_data:
            path_to_place = path_to_place if path_to_place.endswith('\\') else path_to_place + '\\'

            if response_data.status_code == 200:
                print(f'got a response with a status code = {response_data.status_code}')
                with open(path_to_place + str(id_)[:cls._LIMIT_FILE_LEN] + '.xlsx', 'wb') as f:
                    f.write(response_data.content)
                print('xlsx file successfully stored')
            else:
                print(f'Bad request: status code = {response_data.status_code}')

        return cls(path_to_place + str(id_)[:cls._LIMIT_FILE_LEN] + '.xlsx')

    def __init__(self, file_name: str):

        if Path(file_name).is_file():
            self.xlsx_file_name = file_name
        else:
            raise ValueError(f'file_name "{file_name}" is not exists')

    def __enter__(self) -> WrapperXLSX:
        self._xlsx = openpyxl.load_workbook(self.xlsx_file_name)

        # get active sheet
        sheet: Worksheet = self._xlsx.active

        return WrapperXLSX(sheet)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if hasattr(self, '_xlsx'):
            self._xlsx.close()

    def __del__(self):
        # TODO its might be not correct (study the topic)
        self.__exit__(None, None, None)
